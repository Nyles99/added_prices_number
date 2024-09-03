from openpyxl import load_workbook
import os
import csv
import requests
from bs4 import BeautifulSoup


headers = {
    "Accept" : "application/json, text/javascript, */*; q=0.01",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
}

usd_href = "https://www.google.com/search?q=+%D0%BA%D1%83%D1%80%D1%81+%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80%D0%B0&sca_esv=c423d0ff467f72f0&rlz=1C1GCEU_ruRU1112RU1112&sxsrf=ADLYWILLCHXpkziu5DBx2y6869BYtDqAFw%3A1725359298100&ei=wuTWZuPcBYy4wPAPlp_JyAo&ved=0ahUKEwjjxcqgyKaIAxUMHBAIHZZPEqkQ4dUDCA8&oq=+%D0%BA%D1%83%D1%80%D1%81+%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80%D0%B0&gs_lp=Egxnd3Mtd2l6LXNlcnAiGCDQutGD0YDRgSDQtNC-0LvQu9Cw0YDQsDILEAAYgAQYsQMYgwEyCxAAGIAEGLEDGIMBMgsQABiABBixAxiDATILEAAYgAQYsQMYgwEyBRAAGIAEMggQABiABBixAzILEAAYgAQYsQMYgwEyBRAAGIAEMgsQLhiABBjHARivATIFEAAYgARIgAdQAFgAcAB4AZABAJgBNKABNKoBATG4AQzIAQD4AQGYAgGgAj2YAwCSBwExoAeYCg&sclient=gws-wiz-serp"
byn_href = "https://www.google.com/search?q=%D0%BA%D1%83%D1%80%D1%81+byn+%D0%BA+%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80%D1%83&sca_esv=c423d0ff467f72f0&rlz=1C1GCEU_ruRU1112RU1112&sxsrf=ADLYWIKYnhUabzGxH3ft6dCse13GYSBR8A%3A1725361030091&ei=huvWZtimBfe8wPAPreTzmQw&ved=0ahUKEwjYgbvazqaIAxV3HhAIHS3yPMMQ4dUDCA8&uact=5&oq=%D0%BA%D1%83%D1%80%D1%81+byn+%D0%BA+%D0%B4%D0%BE%D0%BB%D0%BB%D0%B0%D1%80%D1%83&gs_lp=Egxnd3Mtd2l6LXNlcnAiHtC60YPRgNGBIGJ5biDQuiDQtNC-0LvQu9Cw0YDRgzIKEAAYgAQYRhiCAjIIEAAYgAQYogQyCBAAGKIEGIkFMggQABiABBiiBDIIEAAYgAQYogQyFhAAGIAEGEYYggIYlwUYjAUY3QTYAQFI8BpQ-A9Y0hZwAngBkAEAmAFPoAGmA6oBATe4AQPIAQD4AQGYAgmgAtoDwgIKEAAYsAMY1gQYR8ICDRAAGIAEGLADGEMYigXCAgUQABiABJgDAIgGAZAGCroGBggBEAEYE5IHATmgB90i&sclient=gws-wiz-serp"
uero_href = "https://www.google.com/search?q=%D0%BA%D1%83%D1%80%D1%81+eur+usd&sca_esv=c423d0ff467f72f0&rlz=1C1GCEU_ruRU1112RU1112&sxsrf=ADLYWIKYnhUabzGxH3ft6dCse13GYSBR8A%3A1725361030091&ei=huvWZtimBfe8wPAPreTzmQw&oq=%D0%BA%D1%83%D1%80%D1%81+uer&gs_lp=Egxnd3Mtd2l6LXNlcnAiDNC60YPRgNGBIHVlcioCCAAyCRAAGIAEGAoYKjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCjIHEAAYgAQYCkjMNFDGI1i9KHADeACQAQCYAUSgAfoBqgEBNLgBAcgBAPgBAZgCBqAC3QHCAgoQABiwAxjWBBhHwgINEAAYgAQYsAMYQxiKBcICCBAAGIAEGLEDwgIKEAAYgAQYFBiHAsICCxAAGIAEGLEDGIMBwgIFEAAYgATCAhEQABiABBixAxgKGCoYRhiCAsICDRAAGIAEGLEDGIMBGArCAh0QABiABBixAxgKGCoYRhiCAhiXBRiMBRjdBNgBAZgDAIgGAZAGCroGBggBEAEYE5IHATagB5wh&sclient=gws-wiz-serp"
req = requests.get(url=usd_href, headers=headers)
src = req.text
soup_1 = BeautifulSoup(src, 'html.parser')
href_part = str(soup_1.find_all("div", class_="b1hJbf"))
usd_rub = float(href_part[href_part.find("data-exchange-rate=")+20 : href_part.find("data-exchange-rate=")+26])
#data-exchange-rate="88.75118">
print(usd_rub, 'Курс доллара к рублю')
req = requests.get(url=byn_href, headers=headers)
src = req.text
soup_1 = BeautifulSoup(src, 'html.parser')
href_part = str(soup_1.find_all("div", class_="b1hJbf"))
usd_byn = float(href_part[href_part.find("data-exchange-rate=")+20 : href_part.find("data-exchange-rate=")+26])
#data-exchange-rate="88.75118">
print(usd_byn, 'Курс доллара к белкам')
req = requests.get(url=uero_href, headers=headers)
src = req.text
soup_1 = BeautifulSoup(src, 'html.parser')
href_part = str(soup_1.find_all("div", class_="b1hJbf"))
usd_uero = float(href_part[href_part.find("data-exchange-rate=")+20 : href_part.find("data-exchange-rate=")+26])
#data-exchange-rate="88.75118">
print(usd_uero, 'Курс доллара к евро')
a = input("Нажми любую кнопку, чтобы закончить")
summa = 1
name_csv = 1

def fix_nulls(s):
    for line in s:
        yield line.replace('\0', '')

def create_file_csv(name_csv):
    if os.path.exists(f"{name_csv}.csv"):
        print("файл csv уже есть")
    else:
        with open(f"{name_csv}.csv", "w", encoding="utf-8") as file_data:
            writer = csv.writer(file_data)

            writer.writerow(
                (
                    'Поставщик',
                    'Артикул',
                    'Ценообразование',
                    'Закупка',
                    'Марка',
                    'Модель',
                    'Год',
                    'Объем двигателя',
                    'Топливо',
                    'Наименование запчасти',
                    'Номера деталей',
                    'Описание',
                    'Фото',
                    'Состояние',
                    'Скидка',
                    'Валюта',
                    'Удаляем дубли',
                    'Цена в долларах',
                    'Цена со скидкой'
                )
            )
create_file_csv(name_csv)
# Указываем путь к директории
directory = "files_csv"
"""spisok_providers = []
book2= load_workbook("provider.xlsx")
sheet2= book2["Лист1"]
for i in range(2, 2000):
    providers = str(sheet2["B"+ str(i)].value)
    spisok_providers.append(providers)"""
# Получаем список файлов
summa_file = 0
files = os.listdir(directory)
for namefile in files:
    summa_file += 1 
# Выводим список файлов
    print(namefile)
    control = 0
    if "-" in namefile:
        provider = ""
        print("KARO")
        text = namefile[namefile.find("by-") +3 : namefile.find(".csv")]
        print(text)
        book= load_workbook("Таблица поставщиков.xlsx")
        sheet = book["Лист1"]
        
        for i in range(2, 2000):
            
            PB_provider = str(sheet["E"+ str(i)].value)
            if text in PB_provider:
                provider = str(sheet["B"+ str(i)].value)
                pricing = str(sheet["C"+ str(i)].value)
                control = 1

        if control != 1:
            if "PB" not in provider:
                pricing = 3
                provider = "PB_&&&"
        
        print(pricing, provider)
        with open(f'files_csv/{namefile}', 'r', encoding="utf-8") as csvfile:
            csvreader = csv.reader(fix_nulls(csvfile))
            n = 1
            p = 1
            
            for row in csvreader:
                
                if p > 1:
                    if summa < 500000 :

                        #print(row)
                        row = str(row)
                        stroka = row.split(";")
                        #for symbol in stroka:
                        #    print(symbol)
                        marka = stroka[1].replace('"',"").replace("'","")
                        model = stroka[2].replace('"',"").replace("'","")
                        year = stroka[3].replace('"',"").replace("'","")
                        volume = stroka[6].replace('"',"").replace("'","")
                        fuel = stroka[5].replace('"',"").replace("'","")
                        name_part = stroka[4].replace('"',"").replace("'","")
                        num_zap = stroka[10].replace('"',"").replace("'","")
                        info = stroka[11].replace('"',"").replace("'","")
                        price = stroka[12].replace('"',"").replace("'","")
                        val = stroka[13].replace('"',"").replace("'","")
                        artical = stroka[0].replace('"',"").replace("['","").replace("'","")
                        foto = stroka[17].replace('"',"").replace("'","")
                        sale = stroka[14].replace('"',"").replace("'","")
                        status = stroka[19].replace('"',"").replace("'","")
                        if val == 'BYN':
                            usd = int(price) * usd_byn
                        elif val == 'RUR':
                            usd = float(int(price) / usd_rub)
                        elif val == 'USD':
                            usd = price
                        elif val == 'EUR':
                            usd = float(int(price) / usd_uero)
                        else:
                            usd = 'Неизвестная валюта'
                        usd_sale = usd
                        if int(sale) > 0:
                            usd_sale = (float(price)/100) * (100-int(sale))
                        else:
                            usd_sale = usd    
                        if status == "1":
                            status = "Новая"
                        else:
                            status = "б/у"

                        file = open(f"{name_csv}.csv", "a", encoding="utf-8", newline='')
                        writer = csv.writer(file)

                        writer.writerow(
                            (
                                provider,
                                artical,
                                pricing,
                                price,
                                marka,
                                model,
                                year,
                                volume,
                                fuel,
                                name_part,
                                num_zap,
                                info,
                                foto,
                                status,
                                sale,
                                val,
                                'Удаляем дубли',
                                usd,
                                usd_sale
                            )
                        )
                        file.close()
                        summa += 1
                    else:
                        summa = 1
                        name_csv += 1
                        create_file_csv(name_csv)

                else:
                    p += 1

    else:
        provider = ""
        print("driver")
        text = namefile[: namefile.find(".csv")]
        print(text)
        book= load_workbook("Таблица поставщиков.xlsx")
        sheet = book["Лист1"]
        
        for i in range(2, 2000):
            
            PB_provider = str(sheet["D"+ str(i)].value)
            if text == PB_provider:
                provider = str(sheet["B"+ str(i)].value)
                pricing = str(sheet["C"+ str(i)].value)
                control = 1

        if control != 1:
            if "PB" not in provider:
                pricing = 3
                provider = "PB_&&&"

            
        print(pricing, provider)
        with open(f'files_csv/{namefile}', 'r', encoding="utf-8") as csvfile:
            csvreader = csv.reader(fix_nulls(csvfile))
            n = 1
            p = 1
            for row in csvreader:
                if p > 1:
                    if summa < 500000 :    
                        #print(row)
                        row = str(row)
                        stroka = row.split(";")
                        #for symbol in stroka:
                        #     print(symbol)
                        
                        marka = stroka[0].replace("['","").replace('"""',"")
                        model = stroka[1].replace('"""',"")
                        year = stroka[2].replace('"""',"")
                        volume = stroka[3].replace('"""',"")
                        fuel = stroka[4].replace('"""',"")
                        name_part = stroka[8].replace('"""',"")
                        num_zap = stroka[9].replace('"""',"")
                        info = stroka[10].replace('"""',"").replace("'","")
                        price = stroka[11].replace('"""',"")
                        val = stroka[12].replace('"""',"")
                        artical = stroka[13].replace('"""',"")
                        foto = stroka[14].replace('"""',"").replace("'","")
                        sale = stroka[15].replace('"""',"")
                        status = stroka[16].replace('"""',"")
                        if status == "1":
                            status = "Новая"
                        else:
                            status = "б/у"
                        #print(provider)
                        #print(marka, model, year, volume, fuel, name_part, num_zap, info, price, val, artical, foto, sale, status, pricing, provider)
                        
                        file = open(f"{name_csv}.csv", "a", encoding="utf-8", newline='')
                        writer = csv.writer(file)

                        writer.writerow(
                            (
                                provider,
                                artical,
                                pricing,
                                price,
                                marka,
                                model,
                                year,
                                volume,
                                fuel,
                                name_part,
                                num_zap,
                                info,
                                foto,
                                status,
                                sale,
                                val,
                                'Удаляем дубли'
                            )
                        )
                        file.close()
                        summa += 1
                    else:
                        summa = 1
                        name_csv += 1
                        create_file_csv(name_csv)
        
                    
                else:
                    p += 1

print(f"Всего обработано {summa_file} файлов")

a = input("Нажми любую кнопку, чтобы закончить")
    
        