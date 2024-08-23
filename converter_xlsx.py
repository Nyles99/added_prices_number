from openpyxl import load_workbook
import os
import csv
import openpyxl


summa = 2
name_csv = 1
summa_file = 0
def fix_nulls(s):
    for line in s:
        yield line.replace('\0', '')

def create_file_xlsx(name_csv, summa_file, summa=2 ):
    if os.path.exists(f"{name_csv}.csv"):
        print("файл csv уже есть")
    else:
        book = openpyxl.Workbook()

        sheet = book.active
        
        sheet['A1'] = 'Поставщик'
        sheet['B1'] = 'Артикул'
        sheet['C1'] = 'Ценообразование'
        sheet['D1'] = 'Закупка'
        sheet['E1'] = 'Марка'
        sheet['F1'] = 'Модель'
        sheet['G1'] = 'Год'
        sheet['H1'] = 'Объем двигателя'
        sheet['I1'] = 'Топливо'
        sheet['J1'] = 'Наименование запчасти'
        sheet['K1'] = 'Номера деталей'
        sheet['L1'] = 'Описание'
        sheet['M1'] = 'Фото'
        sheet['N1'] = 'Состояние'
        sheet['O1'] = 'Скидка'
        sheet['P1'] = 'Валюта'
        sheet['Q1'] = 'Удаляем дубли'

            #book.save(f'{name_csv}.xlsx')
            #book.close()

    #create_file_csv(name_csv)
    # Указываем путь к директории
    directory = "files_csv"
    """spisok_providers = []
    book2= load_workbook("provider.xlsx")
    sheet2= book2["Лист1"]
    for i in range(2, 2000):
        providers = str(sheet2["B"+ str(i)].value)
        spisok_providers.append(providers)"""
    # Получаем список файлов
    
    files = os.listdir(directory)
    for namefile in files:
        summa_file += 1 
    # Выводим список файлов
        print(namefile)
        control = 0
        if "-" in namefile:
            provider = ""
            print("KARO")
            text = namefile[namefile.find("by-") +3 : namefile.find(".csv")]
            print(text)
            book1= load_workbook("Таблица поставщиков.xlsx")
            sheet1 = book1["Лист1"]
            
            for i in range(2, 2000):
                
                PB_provider = str(sheet1["E"+ str(i)].value)
                if text in PB_provider:
                    provider = str(sheet1["B"+ str(i)].value)
                    pricing = str(sheet1["C"+ str(i)].value)
                    control = 1

            if control != 1:
                if "PB" not in provider:
                    pricing = 3
                    provider = "PB_&&&"
            book1.close()
            with open(f'files_csv/{namefile}', 'r', encoding="utf-8") as csvfile:
                csvreader = csv.reader(fix_nulls(csvfile))
                n = 1
                p = 1
                
                for row in csvreader:
                    print(row)
                    if p > 1:
                        if summa < 500000 :

                            #print(row)
                            row = str(row)
                            stroka = row.split(";")
                            #for symbol in stroka:
                            #    print(symbol)
                            marka = stroka[1].replace('"',"").replace("'","")
                            model = stroka[2].replace('"',"").replace("'","")
                            year = stroka[3].replace('"',"").replace("'","")
                            volume = stroka[6].replace('"',"").replace("'","")
                            fuel = stroka[5].replace('"',"").replace("'","")
                            name_part = stroka[4].replace('"',"").replace("'","")
                            num_zap = stroka[10].replace('"',"").replace("'","")
                            info = stroka[11].replace('"',"").replace("'","")
                            price = stroka[12].replace('"',"").replace("'","")
                            val = stroka[13].replace('"',"").replace("'","")
                            artical = stroka[0].replace('"',"").replace("['","").replace("'","")
                            foto = stroka[17].replace('"',"").replace("'","")
                            sale = stroka[14].replace('"',"").replace("'","")
                            status = stroka[19].replace('"',"").replace("'","")
                            if status == "1":
                                status = "Новая"
                            else:
                                status = "б/у"

                            #book = openpyxl.Workbook()

                            #sheet = book.active
                            
                            sheet[f'A{summa}'] = provider
                            sheet[f'B{summa}'] = str(artical)
                            sheet[f'C{summa}'] = pricing
                            sheet[f'D{summa}'] = str(price)
                            sheet[f'E{summa}'] = marka
                            sheet[f'F{summa}'] = str(model)
                            sheet[f'G{summa}'] = year
                            sheet[f'H{summa}'] = volume
                            sheet[f'I{summa}'] = str(fuel)
                            sheet[f'J{summa}'] = name_part
                            sheet[f'K{summa}'] = str(num_zap)
                            sheet[f'L{summa}'] = str(info)
                            sheet[f'M{summa}'] = foto
                            sheet[f'N{summa}'] = status
                            sheet[f'O{summa}'] = sale
                            sheet[f'P{summa}'] = str(val)
                            sheet[f'Q{summa}'] = 'Удаляем дубли'

                            book.save(f'{name_csv}.xlsx')
                            #book.close()                      
                                
                            summa += 1
                        else:
                            summa = 1
                            name_csv += 1
                            #create_file_csv(name_csv)

                    else:
                        p += 1
        else:
            provider = ""
            print("driver")
            text = namefile[: namefile.find(".csv")]
            print(text)
            book1= load_workbook("Таблица поставщиков.xlsx")
            sheet1 = book1["Лист1"]
            
            for i in range(2, 2000):
                
                PB_provider = str(sheet1["D"+ str(i)].value)
                if text == PB_provider:
                    provider = str(sheet1["B"+ str(i)].value)
                    pricing = str(sheet1["C"+ str(i)].value)
                    control = 1

            if control != 1:
                if "PB" not in provider:
                    pricing = 3
                    provider = "PB_&&&"
            book1.close()
            with open(f'files_csv/{namefile}', 'r', encoding="utf-8") as csvfile:
                csvreader = csv.reader(fix_nulls(csvfile))
                n = 1
                p = 1
                for row in csvreader:
                    if p > 1:
                        if summa < 500000 :    
                            #print(row)
                            row = str(row)
                            stroka = row.split(";")
                            #for symbol in stroka:
                            #     print(symbol)
                            
                            marka = stroka[0].replace("['","").replace('"""',"")
                            model = stroka[1].replace('"""',"")
                            year = stroka[2].replace('"""',"")
                            volume = stroka[3].replace('"""',"")
                            fuel = stroka[4].replace('"""',"")
                            name_part = stroka[8].replace('"""',"")
                            num_zap = stroka[9].replace('"""',"")
                            info = stroka[10].replace('"""',"").replace("'","")
                            price = stroka[11].replace('"""',"")
                            val = stroka[12].replace('"""',"")
                            artical = stroka[13].replace('"""',"")
                            foto = stroka[14].replace('"""',"").replace("'","")
                            sale = stroka[15].replace('"""',"")
                            status = stroka[16].replace('"""',"")
                            if status == "1":
                                status = "Новая"
                            else:
                                status = "б/у"
                            #print(provider)
                            #print(marka, model, year, volume, fuel, name_part, num_zap, info, price, val, artical, foto, sale, status, pricing, provider)
                            
                            sheet[f'A{summa}'] = provider
                            sheet[f'B{summa}'] = str(artical)
                            sheet[f'C{summa}'] = pricing
                            sheet[f'D{summa}'] = str(price)
                            sheet[f'E{summa}'] = marka
                            sheet[f'F{summa}'] = str(model)
                            sheet[f'G{summa}'] = year
                            sheet[f'H{summa}'] = volume
                            sheet[f'I{summa}'] = str(fuel)
                            sheet[f'J{summa}'] = name_part
                            sheet[f'K{summa}'] = str(num_zap)
                            sheet[f'L{summa}'] = str(info)
                            sheet[f'M{summa}'] = foto
                            sheet[f'N{summa}'] = status
                            sheet[f'O{summa}'] = sale
                            sheet[f'P{summa}'] = str(val)
                            sheet[f'Q{summa}'] = 'Удаляем дубли'

                            book.save(f'{name_csv}.xlsx')
                            summa += 1
                        else:
                            book.close()
                            name_csv += 1
                            create_file_xlsx(name_csv, summa_file, summa=2)
            
                        
                    else:
                        p += 1

    

print(f"Всего обработано {summa_file} файлов")

a = input("Нажми любую кнопку, чтобы закончить")
    
        